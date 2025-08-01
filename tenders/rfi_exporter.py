# tenders/rfi_exporter.py - Export RFI schedule to Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from .models import TenderAnalysis, RFIItem

class RFIExcelExporter:
    """Export RFI schedule to formatted Excel file"""
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def export_rfi_schedule(self, tender_analysis: TenderAnalysis) -> HttpResponse:
        """Export RFI schedule as Excel file"""
        
        # Create workbook
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "RFI Schedule"
        
        # Set up the worksheet
        self._setup_header(tender_analysis)
        self._setup_column_headers()
        self._populate_rfi_items(tender_analysis)
        self._format_worksheet()
        
        # Save to BytesIO
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"RFI_Schedule_{tender_analysis.project.name.replace(' ', '_')}_{tender_analysis.analysis_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def _setup_header(self, tender_analysis: TenderAnalysis):
        """Set up the header section with project information"""
        
        # Title
        self.ws['A1'] = "REQUEST FOR INFORMATION (RFI) SCHEDULE"
        self.ws['A1'].font = Font(size=16, bold=True)
        
        # Project information
        self.ws['A3'] = "Project:"
        self.ws['B3'] = tender_analysis.project.name
        self.ws['A4'] = "Reference:"
        self.ws['B4'] = tender_analysis.project.reference
        self.ws['A5'] = "Location:"
        self.ws['B5'] = tender_analysis.project.location
        self.ws['A6'] = "Analysis Date:"
        self.ws['B6'] = tender_analysis.analysis_date.strftime('%d/%m/%Y')
        
        # Tender deadline if available
        if tender_analysis.tender_deadline:
            self.ws['A7'] = "Tender Deadline:"
            self.ws['B7'] = tender_analysis.tender_deadline.strftime('%d/%m/%Y %H:%M')
        
        # Bold the labels
        for row in range(3, 8):
            self.ws[f'A{row}'].font = Font(bold=True)
    
    def _setup_column_headers(self):
        """Set up column headers for RFI items"""
        
        headers = [
            ("A", "RFI No.", 10),
            ("B", "Category", 15),
            ("C", "Priority", 12),
            ("D", "Question/Description", 60),
            ("E", "Document Reference", 20),
            ("F", "Location", 20),
            ("G", "Response", 40),
            ("H", "Answered By", 15),
            ("I", "Date Answered", 15),
            ("J", "Status", 12)
        ]
        
        header_row = 10  # Start headers at row 10
        
        for col_letter, header_text, width in headers:
            cell = self.ws[f'{col_letter}{header_row}']
            cell.value = header_text
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column width
            self.ws.column_dimensions[col_letter].width = width
    
    def _populate_rfi_items(self, tender_analysis: TenderAnalysis):
        """Populate the RFI items in the worksheet"""
        
        rfi_items = tender_analysis.rfi_items.all().order_by('priority', 'category')
        start_row = 11  # Data starts at row 11
        
        # Priority colors
        priority_colors = {
            'CRITICAL': 'FFCCCC',  # Light red
            'HIGH': 'FFE6CC',      # Light orange
            'MEDIUM': 'FFFFCC',    # Light yellow
            'LOW': 'E6F3FF'        # Light blue
        }
        
        for idx, rfi_item in enumerate(rfi_items, 1):
            row = start_row + idx - 1
            
            # RFI Number
            self.ws[f'A{row}'] = f"RFI-{idx:03d}"
            
            # Category
            self.ws[f'B{row}'] = rfi_item.category
            
            # Priority
            priority_cell = self.ws[f'C{row}']
            priority_cell.value = rfi_item.priority
            priority_cell.fill = PatternFill(
                start_color=priority_colors.get(rfi_item.priority, 'FFFFFF'),
                end_color=priority_colors.get(rfi_item.priority, 'FFFFFF'),
                fill_type="solid"
            )
            
            # Question
            question_cell = self.ws[f'D{row}']
            question_cell.value = rfi_item.question
            question_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Document Reference
            self.ws[f'E{row}'] = rfi_item.reference
            
            # Location
            self.ws[f'F{row}'] = rfi_item.location_in_document
            
            # Response
            response_cell = self.ws[f'G{row}']
            response_cell.value = rfi_item.answer if rfi_item.is_answered else ""
            response_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Answered By
            self.ws[f'H{row}'] = rfi_item.answered_by if rfi_item.is_answered else ""
            
            # Date Answered
            if rfi_item.answered_date:
                self.ws[f'I{row}'] = rfi_item.answered_date.strftime('%d/%m/%Y')
            
            # Status
            status_cell = self.ws[f'J{row}']
            status_cell.value = "ANSWERED" if rfi_item.is_answered else "PENDING"
            if rfi_item.is_answered:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Set row height for better readability
            self.ws.row_dimensions[row].height = 30
    
    def _format_worksheet(self):
        """Apply final formatting to the worksheet"""
        
        # Add borders to all data cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Find the last row with data
        last_row = self.ws.max_row
        
        # Apply borders to data range
        for row in range(10, last_row + 1):  # From header row to last data row
            for col in range(1, 11):  # Columns A to J
                cell = self.ws.cell(row=row, column=col)
                cell.border = thin_border
        
        # Freeze panes at header row
        self.ws.freeze_panes = 'A11'
        
        # Add summary section
        self._add_summary_section(last_row + 2)
    
    def _add_summary_section(self, start_row: int):
        """Add summary statistics section"""
        
        self.ws[f'A{start_row}'] = "SUMMARY"
        self.ws[f'A{start_row}'].font = Font(size=14, bold=True)
        
        # Count RFI items by priority
        summary_row = start_row + 2
        
        priorities = ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']
        priority_colors = {
            'CRITICAL': 'FFCCCC',
            'HIGH': 'FFE6CC', 
            'MEDIUM': 'FFFFCC',
            'LOW': 'E6F3FF'
        }
        
        self.ws[f'A{summary_row}'] = "Priority"
        self.ws[f'B{summary_row}'] = "Count"
        self.ws[f'C{summary_row}'] = "Answered"
        self.ws[f'D{summary_row}'] = "Pending"
        
        # Bold headers
        for col in ['A', 'B', 'C', 'D']:
            self.ws[f'{col}{summary_row}'].font = Font(bold=True)
        
        for idx, priority in enumerate(priorities):
            row = summary_row + 1 + idx
            
            # Priority name with color
            priority_cell = self.ws[f'A{row}']
            priority_cell.value = priority
            priority_cell.fill = PatternFill(
                start_color=priority_colors[priority],
                end_color=priority_colors[priority],
                fill_type="solid"
            )
            
            # Use COUNTIF formulas to count items
            total_range = f'$C$11:$C${self.ws.max_row}'
            status_range = f'$J$11:$J${self.ws.max_row}'
            
            # Total count
            self.ws[f'B{row}'] = f'=COUNTIF({total_range},"{priority}")'
            
            # Answered count
            self.ws[f'C{row}'] = f'=COUNTIFS({total_range},"{priority}",{status_range},"ANSWERED")'
            
            # Pending count  
            self.ws[f'D{row}'] = f'=COUNTIFS({total_range},"{priority}",{status_range},"PENDING")'


class ContractSummaryExporter:
    """Export contract information summary to Excel"""
    
    def export_contract_summary(self, tender_analysis: TenderAnalysis) -> HttpResponse:
        """Export contract summary as Excel file"""
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contract Summary"
        
        # Title
        ws['A1'] = "CONTRACT INFORMATION SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Project info
        ws['A3'] = "Project:"
        ws['B3'] = tender_analysis.project.name
        ws['A4'] = "Reference:"
        ws['B4'] = tender_analysis.project.reference
        
        row = 6
        
        # Contract Information
        ws[f'A{row}'] = "CONTRACT INFORMATION"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        contract_info = [
            ("Contract Type:", tender_analysis.contract_type or "Not specified"),
            ("Amendments:", ", ".join(tender_analysis.contract_amendments) if tender_analysis.contract_amendments else "None noted"),
        ]
        
        for label, value in contract_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Key Dates
        ws[f'A{row}'] = "KEY DATES"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        date_info = [
            ("Possession Date:", tender_analysis.possession_date.strftime('%d/%m/%Y') if tender_analysis.possession_date else "Not specified"),
            ("Start on Site:", tender_analysis.start_on_site_date.strftime('%d/%m/%Y') if tender_analysis.start_on_site_date else "Not specified"),
            ("Practical Completion:", tender_analysis.practical_completion_date.strftime('%d/%m/%Y') if tender_analysis.practical_completion_date else "Not specified"),
            ("Handover Date:", tender_analysis.handover_date.strftime('%d/%m/%Y') if tender_analysis.handover_date else "Not specified"),
            ("Tender Deadline:", tender_analysis.tender_deadline.strftime('%d/%m/%Y %H:%M') if tender_analysis.tender_deadline else "Not specified"),
        ]
        
        for label, value in date_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Insurance Requirements
        ws[f'A{row}'] = "INSURANCE REQUIREMENTS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        insurance_info = [
            ("Public Liability:", f"£{tender_analysis.public_liability_amount:,.2f}" if tender_analysis.public_liability_amount else "Not specified"),
            ("Employers Liability:", f"£{tender_analysis.employers_liability_amount:,.2f}" if tender_analysis.employers_liability_amount else "Not specified"),
            ("Professional Indemnity:", f"£{tender_analysis.professional_indemnity_amount:,.2f}" if tender_analysis.professional_indemnity_amount else "Not specified"),
            ("Works Insurance:", f"£{tender_analysis.works_insurance_amount:,.2f}" if tender_analysis.works_insurance_amount else "Not specified"),
        ]
        
        for label, value in insurance_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # LADs
        ws[f'A{row}'] = "LIQUIDATED DAMAGES"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        lads_info = []
        if tender_analysis.lads_amount_per_week:
            lads_info.append(("Per Week:", f"£{tender_analysis.lads_amount_per_week:,.2f}"))
        if tender_analysis.lads_amount_per_day:
            lads_info.append(("Per Day:", f"£{tender_analysis.lads_amount_per_day:,.2f}"))
        if tender_analysis.lads_cap_percentage:
            lads_info.append(("Cap (%):", f"{tender_analysis.lads_cap_percentage}%"))
        if tender_analysis.lads_cap_amount:
            lads_info.append(("Cap (Amount):", f"£{tender_analysis.lads_cap_amount:,.2f}"))
        
        if not lads_info:
            lads_info = [("LADs:", "Not specified")]
        
        for label, value in lads_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"Contract_Summary_{tender_analysis.project.name.replace(' ', '_')}_{tender_analysis.analysis_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response